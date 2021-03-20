import json
import openpyxl

workbook = openpyxl.Workbook()
sheet=workbook.active

json_path = "data.json"
excel_path = "data.xlsx"

with open(json_path, "r+", encoding="utf8") as f:
    for (index,line) in enumerate(f):
        print(index)
        line_json = json.loads(line)
        text = line_json["document"][0]["text"]
        qas = line_json["qas"]
        sheet.cell(index+1,1).value = text

        # 跳过不含中心词的数据
        if len(qas) == 0:
            continue
        for i in range(len(qas)):
            qasi = qas[i]
            print(qasi)
            print(qasi[0])
            print(qasi[0]["answers"])
            for j in range(5):
                values = ""
                answers = qasi[j]["answers"]
                for n in answers:
                    values = values + n["text"]
                    values = values + "、"
                sheet.cell(index+1,j+2+5*i).value = values[:-1]
workbook.save(excel_path)